from datetime import date, datetime, time, timedelta
import re
import unicodedata

from flask import Blueprint, jsonify, request, g
from mysql.connector import IntegrityError
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

from db import execute, fetch_all, fetch_one, get_connection
from auth_utils import papel_obrigatorio

admin_bp = Blueprint('admin', __name__)


def texto(v):
    return str(v or '').strip()


def email(v):
    return texto(v).lower()


def digitos(v):
    return ''.join(filter(str.isdigit, str(v or '')))


def normalizar(v):
    valor = unicodedata.normalize('NFKD', texto(v)).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'\s+', ' ', valor).strip().lower()


def erro_integridade(e):
    m = str(e).lower()
    if 'email' in m:
        return 'Este email já está cadastrado.'
    if 'matricula' in m or 'primary' in m:
        return 'Esta matrícula já está cadastrada.'
    return 'Registro duplicado ou vínculo inválido.'


DIAS_PLANILHA = {
    'segunda': 'Segunda',
    'terca': 'Terca',
    'quarta': 'Quarta',
    'quinta': 'Quinta',
    'sexta': 'Sexta',
    'sabado': 'Sabado',
    'domingo': 'Domingo',
}

INTERVALOS = {
    'lanche', 'lanche da manha', 'lanche da tarde', 'intervalo',
    'almoco', 'horario de saida', 'saida', 'recreio'
}


def valor_para_time(valor):
    if valor is None:
        return None
    if isinstance(valor, time):
        return valor.replace(microsecond=0)
    if isinstance(valor, datetime):
        return valor.time().replace(microsecond=0)
    if isinstance(valor, (int, float)):
        segundos = int(round(float(valor) * 24 * 60 * 60)) % (24 * 60 * 60)
        return time(segundos // 3600, (segundos % 3600) // 60, segundos % 60)
    s = texto(valor)
    for formato in ('%H:%M:%S', '%H:%M'):
        try:
            return datetime.strptime(s, formato).time()
        except ValueError:
            pass
    return None


def extrair_grade_planilha(arquivo):
    """Lê o modelo de grade enviado pela escola.

    Linha 1: nomes dos dias.
    Coluna A: horário de início de cada faixa.
    Demais células: matéria ou intervalo.
    O fim de uma aula é o próximo horário existente na coluna A.
    """
    try:
        wb = load_workbook(arquivo, data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError('Não foi possível ler a planilha. Envie um arquivo .xlsx válido.') from exc

    ws = wb.active
    dias_por_coluna = {}
    for col in range(2, ws.max_column + 1):
        dia = DIAS_PLANILHA.get(normalizar(ws.cell(1, col).value))
        if dia:
            dias_por_coluna[col] = dia

    if not dias_por_coluna:
        raise ValueError('A primeira linha da planilha precisa conter os dias da semana.')

    horarios_linhas = {}
    for row in range(2, ws.max_row + 1):
        hora = valor_para_time(ws.cell(row, 1).value)
        if hora:
            horarios_linhas[row] = hora

    if len(horarios_linhas) < 2:
        raise ValueError('A coluna A precisa conter os horários da grade.')

    linhas_com_hora = sorted(horarios_linhas)
    aulas = []
    for indice, row in enumerate(linhas_com_hora[:-1]):
        inicio = horarios_linhas[row]
        fim = horarios_linhas[linhas_com_hora[indice + 1]]
        if fim <= inicio:
            continue

        for col, dia in dias_por_coluna.items():
            materia_original = texto(ws.cell(row, col).value)
            if not materia_original:
                continue
            if normalizar(materia_original) in INTERVALOS:
                continue
            aulas.append({
                'dia': dia,
                'inicio': inicio,
                'fim': fim,
                'materia': re.sub(r'\s+', ' ', materia_original).strip(),
            })

    if not aulas:
        raise ValueError('Nenhuma aula foi encontrada na planilha.')
    return aulas


def materia_id_por_nome(cur, nome):
    cur.execute('SELECT id_materia FROM Materia WHERE nome = %s', (nome,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute('INSERT INTO Materia (nome) VALUES (%s)', (nome,))
    return cur.lastrowid


def professor_unico_da_materia(cur, id_materia):
    cur.execute(
        'SELECT matricula_professor FROM Professor_Materia WHERE id_materia = %s ORDER BY matricula_professor',
        (id_materia,)
    )
    rows = cur.fetchall()
    return rows[0][0] if len(rows) == 1 else None


@admin_bp.route('/resumo', methods=['GET'])
@papel_obrigatorio('administrador')
def resumo():
    return jsonify({
        'professores': fetch_one('''
            SELECT COUNT(*) total FROM Professor p
            JOIN Usuario u ON u.id_usuario = p.id_usuario
            WHERE u.ativo = TRUE
        ''')['total'],
        'coordenadores': fetch_one('''
            SELECT COUNT(*) total FROM Coordenador c
            JOIN Usuario u ON u.id_usuario = c.id_usuario
            WHERE u.ativo = TRUE
        ''')['total'],
        'porteiros': fetch_one('''
            SELECT COUNT(*) total FROM Porteiro p
            JOIN Usuario u ON u.id_usuario = p.id_usuario
            WHERE u.ativo = TRUE
        ''')['total'],
        'alunos': fetch_one('SELECT COUNT(*) total FROM Aluno WHERE ativo = TRUE')['total'],
        'turmas': fetch_one('SELECT COUNT(*) total FROM Turma')['total'],
    })


@admin_bp.route('/materias', methods=['GET'])
@papel_obrigatorio('administrador')
def materias():
    return jsonify(fetch_all('SELECT id_materia, nome FROM Materia WHERE ativo = TRUE ORDER BY nome'))


@admin_bp.route('/turmas', methods=['GET'])
@papel_obrigatorio('administrador')
def turmas():
    return jsonify(fetch_all('''
        SELECT t.codigo, COUNT(h.id_horario) AS total_aulas
        FROM Turma t
        LEFT JOIN Horario h ON h.turma = t.codigo
        GROUP BY t.codigo
        ORDER BY t.codigo
    '''))


@admin_bp.route('/turmas', methods=['POST'])
@papel_obrigatorio('administrador')
def criar_turma():
    codigo = texto(request.form.get('codigo')).upper()
    arquivo = request.files.get('planilha')

    if not codigo or len(codigo) > 10:
        return jsonify({'erro': 'Informe um código de turma com até 10 caracteres.'}), 400
    if not arquivo or not arquivo.filename:
        return jsonify({'erro': 'Selecione a planilha de horários da turma (.xlsx).'}), 400
    if not arquivo.filename.lower().endswith('.xlsx'):
        return jsonify({'erro': 'A grade deve ser enviada em formato .xlsx.'}), 400

    try:
        aulas = extrair_grade_planilha(arquivo.stream)
    except ValueError as exc:
        return jsonify({'erro': str(exc)}), 400

    ano = date.today().year
    inicio_vigencia = date(ano, 1, 1)
    fim_vigencia = date(ano, 12, 31)

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute('INSERT INTO Turma (codigo) VALUES (%s)', (codigo,))
        materias_criadas = set()
        aulas_sem_professor = 0

        for aula in aulas:
            id_materia = materia_id_por_nome(cur, aula['materia'])
            materias_criadas.add(id_materia)
            professor = professor_unico_da_materia(cur, id_materia)
            if professor is None:
                aulas_sem_professor += 1
            cur.execute('''
                INSERT INTO Horario
                (turma, id_materia, matricula_professor, dia_da_semana, hr_inicio, hr_final,
                 data_inicio_vigencia, data_fim_vigencia, id_usuario_cadastro)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                codigo, id_materia, professor, aula['dia'], aula['inicio'], aula['fim'],
                inicio_vigencia, fim_vigencia, g.usuario['id_usuario']
            ))

        conn.commit()
        return jsonify({
            'mensagem': 'Turma e grade de horários importadas com sucesso.',
            'aulas_importadas': len(aulas),
            'materias_identificadas': len(materias_criadas),
            'aulas_sem_professor': aulas_sem_professor,
        }), 201
    except IntegrityError as exc:
        conn.rollback()
        if 'turma.primary' in str(exc).lower() or 'duplicate' in str(exc).lower():
            return jsonify({'erro': 'Já existe uma turma com esse código.'}), 409
        return jsonify({'erro': erro_integridade(exc)}), 409
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/alunos', methods=['GET'])
@papel_obrigatorio('administrador')
def alunos():
    incluir_inativos = request.args.get('incluir_inativos') == '1'
    filtro = '' if incluir_inativos else 'WHERE a.ativo = TRUE'
    return jsonify(fetch_all(f'''
        SELECT a.matricula, a.nome, a.turma, r.nome AS responsavel, a.ativo
        FROM Aluno a
        LEFT JOIN Responsavel r ON r.id_responsavel = a.id_responsavel
        {filtro}
        ORDER BY a.ativo DESC, a.nome
    '''))


@admin_bp.route('/alunos', methods=['POST'])
@papel_obrigatorio('administrador')
def criar_aluno():
    dados = request.get_json() or {}
    nome = texto(dados.get('nome'))
    turma = texto(dados.get('turma')).upper()
    matricula = texto(dados.get('matricula'))
    if not nome or not turma or not matricula:
        return jsonify({'erro': 'Nome, turma e matrícula são obrigatórios.'}), 400
    if not matricula.isdigit() or not 6 <= len(matricula) <= 12:
        return jsonify({'erro': 'A matrícula deve conter entre 6 e 12 dígitos.'}), 400

    existente = fetch_one('SELECT ativo FROM Aluno WHERE matricula = %s', (matricula,))
    if existente:
        if existente['ativo']:
            return jsonify({'erro': 'Esta matrícula já está cadastrada.'}), 409
        return jsonify({'erro': 'Esta matrícula pertence a um aluno desativado. Use o botão Reativar na lista em vez de criar outro cadastro.'}), 409

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute('INSERT INTO Aluno (matricula, nome, turma) VALUES (%s, %s, %s)', (matricula, nome, turma))
        conn.commit()
        return jsonify({'mensagem': 'Aluno cadastrado.', 'matricula': matricula}), 201
    except IntegrityError as erro:
        conn.rollback()
        return jsonify({'erro': erro_integridade(erro)}), 409
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/alunos/<matricula>', methods=['DELETE'])
@papel_obrigatorio('administrador')
def excluir_aluno(matricula):
    aluno = fetch_one('SELECT matricula, ativo FROM Aluno WHERE matricula = %s', (matricula,))
    if not aluno:
        return jsonify({'erro': 'Aluno não encontrado.'}), 404

    permanente = request.args.get('permanente') == '1'
    if not permanente:
        if not aluno['ativo']:
            return jsonify({'erro': 'Este aluno já está desativado.'}), 409
        execute('UPDATE Aluno SET ativo = FALSE WHERE matricula = %s', (matricula,))
        return jsonify({'mensagem': 'Aluno desativado. O histórico foi preservado.'})

    if aluno['ativo']:
        return jsonify({'erro': 'Desative o aluno antes de solicitar a exclusão permanente.'}), 409

    historico = fetch_one('SELECT COUNT(*) total FROM Ocorrencia_Aluno WHERE matricula = %s', (matricula,))['total']
    if historico:
        return jsonify({
            'erro': 'Este aluno possui ocorrências/histórico e não pode ser excluído definitivamente. Mantenha-o desativado.'
        }), 409
    execute('DELETE FROM Aluno WHERE matricula = %s', (matricula,))
    return jsonify({'mensagem': 'Aluno excluído definitivamente. Nenhum histórico foi removido.'})


@admin_bp.route('/alunos/<matricula>/reativar', methods=['POST'])
@papel_obrigatorio('administrador')
def reativar_aluno(matricula):
    aluno = fetch_one('SELECT matricula, ativo FROM Aluno WHERE matricula = %s', (matricula,))
    if not aluno:
        return jsonify({'erro': 'Aluno não encontrado.'}), 404
    if aluno['ativo']:
        return jsonify({'erro': 'Este aluno já está ativo.'}), 409
    execute('UPDATE Aluno SET ativo = TRUE WHERE matricula = %s', (matricula,))
    return jsonify({'mensagem': 'Aluno reativado com sucesso.'})


def pessoa_com_usuario(tabela, coluna_id, valor_id):
    return fetch_one(f'''
        SELECT p.{coluna_id} AS id_pessoa, p.id_usuario, u.ativo
        FROM {tabela} p
        JOIN Usuario u ON u.id_usuario = p.id_usuario
        WHERE p.{coluna_id} = %s
    ''', (valor_id,))


def desativar_usuario_da_pessoa(tabela, coluna_id, valor_id):
    """Desativa o acesso sem apagar a pessoa nem seu histórico."""
    pessoa = pessoa_com_usuario(tabela, coluna_id, valor_id)
    if not pessoa:
        return False, 'nao_encontrado'
    if not pessoa['ativo']:
        return False, 'ja_inativo'
    execute('UPDATE Usuario SET ativo = FALSE WHERE id_usuario = %s', (pessoa['id_usuario'],))
    return True, None


def reativar_usuario_da_pessoa(tabela, coluna_id, valor_id):
    pessoa = pessoa_com_usuario(tabela, coluna_id, valor_id)
    if not pessoa:
        return False, 'nao_encontrado'
    if pessoa['ativo']:
        return False, 'ja_ativo'
    execute('UPDATE Usuario SET ativo = TRUE WHERE id_usuario = %s', (pessoa['id_usuario'],))
    return True, None


def excluir_usuario_definitivamente(tabela, coluna_id, valor_id, tipo):
    """Exclui somente cadastros sem histórico importante.

    A exclusão permanente é propositalmente conservadora: quando houver algo que
    represente histórico escolar, a API bloqueia e orienta manter o cadastro inativo.
    """
    pessoa = pessoa_com_usuario(tabela, coluna_id, valor_id)
    if not pessoa:
        return None, (jsonify({'erro': 'Registro não encontrado.'}), 404)

    id_usuario = pessoa['id_usuario']
    if pessoa['ativo']:
        return None, (jsonify({'erro': 'Desative este cadastro antes de solicitar a exclusão permanente.'}), 409)

    if tipo == 'professor':
        historico_aulas = fetch_one(
            'SELECT COUNT(*) total FROM Ocorrencia_Aula WHERE matricula_professor = %s',
            (valor_id,)
        )['total']
        historico_ocorrencias = fetch_one(
            'SELECT COUNT(*) total FROM Ocorrencia WHERE matricula_professor = %s',
            (valor_id,)
        )['total']
        if historico_aulas or historico_ocorrencias:
            return None, (jsonify({
                'erro': 'Este professor possui histórico de aulas/ocorrências e não pode ser excluído definitivamente. Mantenha-o desativado.'
            }), 409)
    elif tipo == 'coordenador':
        aprovacoes = fetch_one(
            'SELECT COUNT(*) total FROM Ocorrencia WHERE id_usuario_aprovador = %s',
            (id_usuario,)
        )['total']
        if aprovacoes:
            return None, (jsonify({
                'erro': 'Este coordenador possui decisões registradas na Gestão e não pode ser excluído definitivamente. Mantenha-o desativado.'
            }), 409)

    conn = get_connection()
    cur = conn.cursor()
    try:
        if tipo == 'professor':
            # Horários são preservados; apenas deixam de apontar para o cadastro removido.
            cur.execute('UPDATE Horario SET matricula_professor = NULL WHERE matricula_professor = %s', (valor_id,))
        cur.execute(f'DELETE FROM {tabela} WHERE {coluna_id} = %s', (valor_id,))
        cur.execute('DELETE FROM Usuario WHERE id_usuario = %s', (id_usuario,))
        conn.commit()
        return {'mensagem': 'Cadastro excluído definitivamente. Nenhum histórico importante foi removido.'}, None
    except IntegrityError:
        conn.rollback()
        return None, (jsonify({
            'erro': 'Este cadastro ainda possui vínculos no sistema e não pode ser excluído definitivamente. Desative-o para preservar os dados.'
        }), 409)
    finally:
        cur.close()
        conn.close()


def criar_pessoa(dados, nivel_acesso, tabela):
    nome = texto(dados.get('nome'))
    email_normalizado = email(dados.get('email'))
    senha = str(dados.get('senha') or '')
    telefone = digitos(dados.get('telefone')) or None

    if not nome or not email_normalizado or not senha:
        return None, (jsonify({'erro': 'Nome, email e senha são obrigatórios.'}), 400)
    if len(senha) < 8:
        return None, (jsonify({'erro': 'A senha inicial deve ter pelo menos 8 caracteres.'}), 400)
    if telefone and len(telefone) not in (10, 11):
        return None, (jsonify({'erro': 'O telefone deve conter 10 ou 11 dígitos.'}), 400)

    usuario_existente = fetch_one('SELECT ativo FROM Usuario WHERE email = %s', (email_normalizado,))
    if usuario_existente:
        if usuario_existente['ativo']:
            return None, (jsonify({'erro': 'Este email já está cadastrado.'}), 409)
        return None, (jsonify({'erro': 'Este email pertence a um usuário desativado. Reative o cadastro existente ou exclua-o definitivamente se for um cadastro feito por engano.'}), 409)

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            'INSERT INTO Usuario (email, senha, telefone, nivel_acesso) VALUES (%s, %s, %s, %s)',
            (email_normalizado, generate_password_hash(senha), telefone, nivel_acesso)
        )
        id_usuario = cur.lastrowid
        cur.execute(f'INSERT INTO {tabela} (id_usuario, nome) VALUES (%s, %s)', (id_usuario, nome))
        id_pessoa = cur.lastrowid
        conn.commit()
        return {'mensagem': 'Usuário cadastrado.', 'id_pessoa': id_pessoa}, None
    except IntegrityError as erro:
        conn.rollback()
        return None, (jsonify({'erro': erro_integridade(erro)}), 409)
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/professores', methods=['GET'])
@papel_obrigatorio('administrador')
def professores():
    incluir_inativos = request.args.get('incluir_inativos') == '1'
    filtro = '' if incluir_inativos else 'WHERE u.ativo = TRUE'
    return jsonify(fetch_all(f'''
        SELECT p.matricula, p.nome, u.email, u.telefone, u.ativo,
               MIN(pm.id_materia) AS id_materia,
               GROUP_CONCAT(DISTINCT pm.id_materia ORDER BY pm.id_materia SEPARATOR ',') AS materias_ids,
               GROUP_CONCAT(DISTINCT m.nome ORDER BY m.nome SEPARATOR ', ') AS materias
        FROM Professor p
        JOIN Usuario u ON u.id_usuario = p.id_usuario
        LEFT JOIN Professor_Materia pm ON pm.matricula_professor = p.matricula
        LEFT JOIN Materia m ON m.id_materia = pm.id_materia
        {filtro}
        GROUP BY p.matricula, p.nome, u.email, u.telefone, u.ativo
        ORDER BY u.ativo DESC, p.nome
    '''))


@admin_bp.route('/professores', methods=['POST'])
@papel_obrigatorio('administrador')
def criar_professor():
    dados = request.get_json() or {}
    nome = texto(dados.get('nome'))
    email_normalizado = email(dados.get('email'))
    senha = str(dados.get('senha') or '')
    telefone = digitos(dados.get('telefone')) or None
    try:
        id_materia = int(dados.get('id_materia'))
    except (TypeError, ValueError):
        return jsonify({'erro': 'Selecione a matéria do professor.'}), 400

    if not nome or not email_normalizado or not senha:
        return jsonify({'erro': 'Nome, email e senha são obrigatórios.'}), 400
    if len(senha) < 8:
        return jsonify({'erro': 'A senha inicial deve ter pelo menos 8 caracteres.'}), 400
    if telefone and len(telefone) not in (10, 11):
        return jsonify({'erro': 'O telefone deve conter 10 ou 11 dígitos.'}), 400

    usuario_existente = fetch_one('SELECT ativo FROM Usuario WHERE email = %s', (email_normalizado,))
    if usuario_existente:
        if usuario_existente['ativo']:
            return jsonify({'erro': 'Este email já está cadastrado.'}), 409
        return jsonify({'erro': 'Este email pertence a um professor desativado. Reative o cadastro existente ou exclua-o definitivamente se tiver sido criado por engano.'}), 409

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute('SELECT id_materia FROM Materia WHERE id_materia = %s AND ativo = TRUE', (id_materia,))
        if not cur.fetchone():
            conn.rollback()
            return jsonify({'erro': 'Matéria não encontrada.'}), 404

        cur.execute('''
            INSERT INTO Usuario (email, senha, telefone, nivel_acesso)
            VALUES (%s, %s, %s, 2)
        ''', (email_normalizado, generate_password_hash(senha), telefone))
        id_usuario = cur.lastrowid
        cur.execute('INSERT INTO Professor (id_usuario, nome) VALUES (%s, %s)', (id_usuario, nome))
        matricula_professor = cur.lastrowid
        cur.execute('''
            INSERT INTO Professor_Materia (matricula_professor, id_materia)
            VALUES (%s, %s)
        ''', (matricula_professor, id_materia))

        cur.execute('SELECT COUNT(*) FROM Professor_Materia WHERE id_materia = %s', (id_materia,))
        total_professores_materia = cur.fetchone()[0]
        aulas_vinculadas = 0
        if total_professores_materia == 1:
            cur.execute('''
                UPDATE Horario SET matricula_professor = %s
                WHERE id_materia = %s AND matricula_professor IS NULL
            ''', (matricula_professor, id_materia))
            aulas_vinculadas = cur.rowcount

        conn.commit()
        return jsonify({
            'mensagem': 'Professor cadastrado e vinculado à matéria.',
            'aulas_vinculadas': aulas_vinculadas,
        }), 201
    except IntegrityError as erro:
        conn.rollback()
        return jsonify({'erro': erro_integridade(erro)}), 409
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/professores/<int:matricula>', methods=['DELETE'])
@papel_obrigatorio('administrador')
def excluir_professor(matricula):
    if request.args.get('permanente') == '1':
        resultado, erro = excluir_usuario_definitivamente('Professor', 'matricula', matricula, 'professor')
        return erro if erro else jsonify(resultado)
    ok, motivo = desativar_usuario_da_pessoa('Professor', 'matricula', matricula)
    if motivo == 'nao_encontrado':
        return jsonify({'erro': 'Professor não encontrado.'}), 404
    if motivo == 'ja_inativo':
        return jsonify({'erro': 'Este professor já está desativado.'}), 409
    return jsonify({'mensagem': 'Professor desativado. O histórico foi preservado.'})


@admin_bp.route('/professores/<int:matricula>/reativar', methods=['POST'])
@papel_obrigatorio('administrador')
def reativar_professor(matricula):
    ok, motivo = reativar_usuario_da_pessoa('Professor', 'matricula', matricula)
    if motivo == 'nao_encontrado':
        return jsonify({'erro': 'Professor não encontrado.'}), 404
    if motivo == 'ja_ativo':
        return jsonify({'erro': 'Este professor já está ativo.'}), 409
    return jsonify({'mensagem': 'Professor reativado com sucesso.'})


@admin_bp.route('/coordenadores', methods=['GET'])
@papel_obrigatorio('administrador')
def coordenadores():
    incluir_inativos = request.args.get('incluir_inativos') == '1'
    filtro = '' if incluir_inativos else 'WHERE u.ativo = TRUE'
    return jsonify(fetch_all(f'''
        SELECT c.id_coordenador AS id, c.nome, u.email, u.telefone, u.ativo
        FROM Coordenador c JOIN Usuario u ON u.id_usuario = c.id_usuario
        {filtro}
        ORDER BY u.ativo DESC, c.nome
    '''))


@admin_bp.route('/coordenadores', methods=['POST'])
@papel_obrigatorio('administrador')
def criar_coordenador():
    resultado, erro = criar_pessoa(request.get_json() or {}, 3, 'Coordenador')
    return erro if erro else (jsonify({'mensagem': resultado['mensagem']}), 201)


@admin_bp.route('/coordenadores/<int:id_coordenador>', methods=['DELETE'])
@papel_obrigatorio('administrador')
def excluir_coordenador(id_coordenador):
    if request.args.get('permanente') == '1':
        resultado, erro = excluir_usuario_definitivamente('Coordenador', 'id_coordenador', id_coordenador, 'coordenador')
        return erro if erro else jsonify(resultado)
    ok, motivo = desativar_usuario_da_pessoa('Coordenador', 'id_coordenador', id_coordenador)
    if motivo == 'nao_encontrado':
        return jsonify({'erro': 'Coordenador(a) não encontrado(a).'}), 404
    if motivo == 'ja_inativo':
        return jsonify({'erro': 'Este coordenador já está desativado.'}), 409
    return jsonify({'mensagem': 'Coordenador(a) desativado(a). O histórico foi preservado.'})


@admin_bp.route('/coordenadores/<int:id_coordenador>/reativar', methods=['POST'])
@papel_obrigatorio('administrador')
def reativar_coordenador(id_coordenador):
    ok, motivo = reativar_usuario_da_pessoa('Coordenador', 'id_coordenador', id_coordenador)
    if motivo == 'nao_encontrado':
        return jsonify({'erro': 'Coordenador(a) não encontrado(a).'}), 404
    if motivo == 'ja_ativo':
        return jsonify({'erro': 'Este coordenador já está ativo.'}), 409
    return jsonify({'mensagem': 'Coordenador(a) reativado(a) com sucesso.'})


@admin_bp.route('/porteiros', methods=['GET'])
@papel_obrigatorio('administrador')
def porteiros():
    incluir_inativos = request.args.get('incluir_inativos') == '1'
    filtro = '' if incluir_inativos else 'WHERE u.ativo = TRUE'
    return jsonify(fetch_all(f'''
        SELECT p.id_porteiro AS id, p.nome, u.email, u.telefone, u.ativo
        FROM Porteiro p JOIN Usuario u ON u.id_usuario = p.id_usuario
        {filtro}
        ORDER BY u.ativo DESC, p.nome
    '''))


@admin_bp.route('/porteiros', methods=['POST'])
@papel_obrigatorio('administrador')
def criar_porteiro():
    resultado, erro = criar_pessoa(request.get_json() or {}, 4, 'Porteiro')
    return erro if erro else (jsonify({'mensagem': resultado['mensagem']}), 201)


@admin_bp.route('/porteiros/<int:id_porteiro>', methods=['DELETE'])
@papel_obrigatorio('administrador')
def excluir_porteiro(id_porteiro):
    if request.args.get('permanente') == '1':
        resultado, erro = excluir_usuario_definitivamente('Porteiro', 'id_porteiro', id_porteiro, 'porteiro')
        return erro if erro else jsonify(resultado)
    ok, motivo = desativar_usuario_da_pessoa('Porteiro', 'id_porteiro', id_porteiro)
    if motivo == 'nao_encontrado':
        return jsonify({'erro': 'Porteiro(a) não encontrado(a).'}), 404
    if motivo == 'ja_inativo':
        return jsonify({'erro': 'Este porteiro já está desativado.'}), 409
    return jsonify({'mensagem': 'Porteiro(a) desativado(a). O histórico foi preservado.'})


@admin_bp.route('/porteiros/<int:id_porteiro>/reativar', methods=['POST'])
@papel_obrigatorio('administrador')
def reativar_porteiro(id_porteiro):
    ok, motivo = reativar_usuario_da_pessoa('Porteiro', 'id_porteiro', id_porteiro)
    if motivo == 'nao_encontrado':
        return jsonify({'erro': 'Porteiro(a) não encontrado(a).'}), 404
    if motivo == 'ja_ativo':
        return jsonify({'erro': 'Este porteiro já está ativo.'}), 409
    return jsonify({'mensagem': 'Porteiro(a) reativado(a) com sucesso.'})


def validar_email_disponivel(cur, email_normalizado, id_usuario_atual):
    cur.execute('SELECT id_usuario FROM Usuario WHERE email = %s AND id_usuario <> %s', (email_normalizado, id_usuario_atual))
    return cur.fetchone() is None


@admin_bp.route('/alunos/<matricula>', methods=['PUT'])
@papel_obrigatorio('administrador')
def editar_aluno(matricula):
    dados = request.get_json() or {}
    nova_matricula = texto(dados.get('matricula'))
    nome = texto(dados.get('nome'))
    turma = texto(dados.get('turma')).upper()
    if not nome or not turma or not nova_matricula:
        return jsonify({'erro': 'Nome, turma e matrícula são obrigatórios.'}), 400
    if not nova_matricula.isdigit() or not 6 <= len(nova_matricula) <= 12:
        return jsonify({'erro': 'A matrícula deve conter entre 6 e 12 dígitos.'}), 400
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute('SELECT matricula FROM Aluno WHERE matricula = %s', (matricula,))
        if not cur.fetchone(): return jsonify({'erro': 'Aluno não encontrado.'}), 404
        cur.execute('SELECT codigo FROM Turma WHERE codigo = %s', (turma,))
        if not cur.fetchone(): return jsonify({'erro': 'Turma não encontrada.'}), 404
        if nova_matricula != matricula:
            cur.execute('SELECT matricula FROM Aluno WHERE matricula = %s', (nova_matricula,))
            if cur.fetchone(): return jsonify({'erro': 'A nova matrícula já está cadastrada.'}), 409
        cur.execute('UPDATE Aluno SET matricula=%s, nome=%s, turma=%s WHERE matricula=%s',
                    (nova_matricula, nome, turma, matricula))
        conn.commit()
        return jsonify({'mensagem': 'Aluno atualizado com sucesso.', 'matricula': nova_matricula})
    except IntegrityError as exc:
        conn.rollback(); return jsonify({'erro': erro_integridade(exc)}), 409
    finally:
        cur.close(); conn.close()


@admin_bp.route('/professores/<int:matricula>', methods=['PUT'])
@papel_obrigatorio('administrador')
def editar_professor(matricula):
    dados = request.get_json() or {}
    nome = texto(dados.get('nome')); email_normalizado = email(dados.get('email'))
    telefone = digitos(dados.get('telefone')) or None
    try: id_materia = int(dados.get('id_materia'))
    except (TypeError, ValueError): return jsonify({'erro': 'Selecione a matéria do professor.'}), 400
    if not nome or not email_normalizado: return jsonify({'erro': 'Nome e email são obrigatórios.'}), 400
    if telefone and len(telefone) not in (10, 11): return jsonify({'erro': 'O telefone deve conter 10 ou 11 dígitos.'}), 400
    conn=get_connection(); cur=conn.cursor()
    try:
        cur.execute('SELECT id_usuario FROM Professor WHERE matricula=%s', (matricula,)); row=cur.fetchone()
        if not row: return jsonify({'erro': 'Professor não encontrado.'}), 404
        id_usuario=row[0]
        if not validar_email_disponivel(cur, email_normalizado, id_usuario): return jsonify({'erro':'Este email já está cadastrado.'}),409
        cur.execute('SELECT id_materia FROM Materia WHERE id_materia=%s AND ativo=TRUE',(id_materia,))
        if not cur.fetchone(): return jsonify({'erro':'Matéria não encontrada.'}),404
        cur.execute('UPDATE Professor SET nome=%s WHERE matricula=%s',(nome,matricula))
        cur.execute('UPDATE Usuario SET email=%s, telefone=%s WHERE id_usuario=%s',(email_normalizado,telefone,id_usuario))
        cur.execute('UPDATE Horario SET matricula_professor=NULL WHERE matricula_professor=%s AND id_materia<>%s',(matricula,id_materia))
        cur.execute('DELETE FROM Professor_Materia WHERE matricula_professor=%s',(matricula,))
        cur.execute('INSERT INTO Professor_Materia (matricula_professor,id_materia) VALUES (%s,%s)',(matricula,id_materia))
        cur.execute('SELECT COUNT(*) FROM Professor_Materia WHERE id_materia=%s',(id_materia,)); total=cur.fetchone()[0]
        if total == 1:
            cur.execute('UPDATE Horario SET matricula_professor=%s WHERE id_materia=%s AND matricula_professor IS NULL',(matricula,id_materia))
        conn.commit(); return jsonify({'mensagem':'Professor atualizado com sucesso.'})
    except IntegrityError as exc:
        conn.rollback(); return jsonify({'erro':erro_integridade(exc)}),409
    finally: cur.close(); conn.close()


def editar_pessoa_simples(tabela, coluna_id, valor_id):
    dados=request.get_json() or {}; nome=texto(dados.get('nome')); email_normalizado=email(dados.get('email')); telefone=digitos(dados.get('telefone')) or None
    if not nome or not email_normalizado: return None,(jsonify({'erro':'Nome e email são obrigatórios.'}),400)
    if telefone and len(telefone) not in (10,11): return None,(jsonify({'erro':'O telefone deve conter 10 ou 11 dígitos.'}),400)
    conn=get_connection(); cur=conn.cursor()
    try:
        cur.execute(f'SELECT id_usuario FROM {tabela} WHERE {coluna_id}=%s',(valor_id,)); row=cur.fetchone()
        if not row: return None,(jsonify({'erro':'Registro não encontrado.'}),404)
        id_usuario=row[0]
        if not validar_email_disponivel(cur,email_normalizado,id_usuario): return None,(jsonify({'erro':'Este email já está cadastrado.'}),409)
        cur.execute(f'UPDATE {tabela} SET nome=%s WHERE {coluna_id}=%s',(nome,valor_id))
        cur.execute('UPDATE Usuario SET email=%s, telefone=%s WHERE id_usuario=%s',(email_normalizado,telefone,id_usuario))
        conn.commit(); return {'mensagem':'Cadastro atualizado com sucesso.'},None
    except IntegrityError as exc:
        conn.rollback(); return None,(jsonify({'erro':erro_integridade(exc)}),409)
    finally: cur.close(); conn.close()


@admin_bp.route('/coordenadores/<int:id_coordenador>', methods=['PUT'])
@papel_obrigatorio('administrador')
def editar_coordenador(id_coordenador):
    resultado,erro=editar_pessoa_simples('Coordenador','id_coordenador',id_coordenador)
    return erro if erro else jsonify(resultado)


@admin_bp.route('/porteiros/<int:id_porteiro>', methods=['PUT'])
@papel_obrigatorio('administrador')
def editar_porteiro(id_porteiro):
    resultado,erro=editar_pessoa_simples('Porteiro','id_porteiro',id_porteiro)
    return erro if erro else jsonify(resultado)

@admin_bp.route('/horarios/<int:id_horario>/materia', methods=['PUT'])
@papel_obrigatorio('administrador')
def editar_materia_horario(id_horario):
    """Corrige a matéria de uma aula específica, inclusive em grades históricas."""
    dados = request.get_json() or {}
    nome = re.sub(r'\s+', ' ', texto(dados.get('nome'))).strip()
    if not nome:
        return jsonify({'erro': 'Informe o nome da matéria.'}), 400
    if len(nome) > 100:
        return jsonify({'erro': 'O nome da matéria deve ter até 100 caracteres.'}), 400
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute('SELECT id_horario, matricula_professor FROM Horario WHERE id_horario=%s', (id_horario,))
        horario = cur.fetchone()
        if not horario:
            return jsonify({'erro': 'Horário não encontrado.'}), 404
        nome_normalizado = normalizar(nome)
        cur.execute('SELECT id_materia, nome FROM Materia WHERE ativo=TRUE')
        materia_existente = next((row for row in cur.fetchall() if normalizar(row[1]) == nome_normalizado), None)
        if materia_existente:
            id_materia, nome_final = materia_existente
        else:
            cur.execute('INSERT INTO Materia (nome) VALUES (%s)', (nome,))
            id_materia, nome_final = cur.lastrowid, nome
        professor = horario[1]
        if professor is not None:
            cur.execute('SELECT 1 FROM Professor_Materia WHERE matricula_professor=%s AND id_materia=%s', (professor, id_materia))
            if not cur.fetchone():
                professor = None
        cur.execute('UPDATE Horario SET id_materia=%s, matricula_professor=%s WHERE id_horario=%s', (id_materia, professor, id_horario))
        conn.commit()
        return jsonify({'mensagem': 'Matéria da aula atualizada com sucesso.', 'id_materia': id_materia, 'materia': nome_final, 'matricula_professor': professor})
    except IntegrityError as exc:
        conn.rollback(); return jsonify({'erro': erro_integridade(exc)}), 409
    finally:
        cur.close(); conn.close()



@admin_bp.route('/responsaveis', methods=['GET'])
@papel_obrigatorio('administrador')
def responsaveis():
    incluir_inativos = request.args.get('incluir_inativos') == '1'
    filtro = '' if incluir_inativos else 'WHERE u.ativo = TRUE'
    return jsonify(fetch_all(f'''
        SELECT r.id_responsavel, r.id_usuario, r.cpf, r.nome, r.telefone,
               u.email, u.ativo,
               GROUP_CONCAT(DISTINCT CONCAT(a.matricula, ' — ', a.nome) ORDER BY a.nome SEPARATOR ' | ') AS alunos
        FROM Responsavel r
        JOIN Usuario u ON u.id_usuario = r.id_usuario
        LEFT JOIN Aluno a ON a.id_responsavel = r.id_responsavel
        {filtro}
        GROUP BY r.id_responsavel, r.id_usuario, r.cpf, r.nome, r.telefone, u.email, u.ativo
        ORDER BY u.ativo DESC, r.nome
    '''))


@admin_bp.route('/responsaveis', methods=['POST'])
@papel_obrigatorio('administrador')
def criar_responsavel_admin():
    dados = request.get_json() or {}
    nome = texto(dados.get('nome'))
    cpf = digitos(dados.get('cpf'))
    telefone = digitos(dados.get('telefone'))
    email_normalizado = email(dados.get('email'))
    senha = str(dados.get('senha') or '')
    matricula = texto(dados.get('matricula'))

    if not all((nome, cpf, telefone, email_normalizado, senha, matricula)):
        return jsonify({'erro': 'Nome, CPF, telefone, email, senha e matrícula são obrigatórios.'}), 400
    if not matricula.isdigit() or not 6 <= len(matricula) <= 12:
        return jsonify({'erro': 'A matrícula deve conter entre 6 e 12 dígitos.'}), 400
    if len(cpf) != 11:
        return jsonify({'erro': 'O CPF deve possuir 11 números.'}), 400
    if len(telefone) not in (10, 11):
        return jsonify({'erro': 'O telefone deve conter 10 ou 11 dígitos.'}), 400
    if len(senha) < 6:
        return jsonify({'erro': 'A senha deve possuir pelo menos 6 caracteres.'}), 400
    if '@' not in email_normalizado:
        return jsonify({'erro': 'Informe um email válido.'}), 400

    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute('SELECT matricula, nome, id_responsavel FROM Aluno WHERE matricula = %s FOR UPDATE', (matricula,))
        aluno = cur.fetchone()
        if not aluno:
            return jsonify({'erro': 'Matrícula não encontrada.'}), 404
        if aluno['id_responsavel'] is not None:
            return jsonify({'erro': 'Esta matrícula já está vinculada a um responsável.'}), 409
        cur.execute('SELECT id_usuario FROM Usuario WHERE email = %s', (email_normalizado,))
        if cur.fetchone():
            return jsonify({'erro': 'Este email já está cadastrado.'}), 409
        cur.execute('SELECT id_responsavel FROM Responsavel WHERE cpf = %s', (cpf,))
        if cur.fetchone():
            return jsonify({'erro': 'Este CPF já está cadastrado.'}), 409
        senha_hash = generate_password_hash(senha)
        cur.execute('INSERT INTO Usuario (email, senha, telefone, nivel_acesso) VALUES (%s, %s, %s, 1)', (email_normalizado, senha_hash, telefone))
        id_usuario = cur.lastrowid
        cur.execute('INSERT INTO Responsavel (id_usuario, cpf, nome, telefone, primeiro_login) VALUES (%s, %s, %s, %s, FALSE)', (id_usuario, cpf, nome, telefone))
        id_responsavel = cur.lastrowid
        cur.execute('UPDATE Aluno SET id_responsavel = %s WHERE matricula = %s', (id_responsavel, matricula))
        conn.commit()
        return jsonify({'mensagem': f'Responsável cadastrado e vinculado a {aluno["nome"]} com sucesso.'}), 201
    except IntegrityError as exc:
        conn.rollback(); return jsonify({'erro': erro_integridade(exc)}), 409
    except Exception as exc:
        conn.rollback(); print('Erro ao cadastrar responsável pelo administrador:', exc)
        return jsonify({'erro': 'Erro interno ao cadastrar responsável.'}), 500
    finally:
        cur.close(); conn.close()


@admin_bp.route('/ocorrencias', methods=['GET'])
@papel_obrigatorio('administrador')
def historico_ocorrencias_admin():
    categoria = texto(request.args.get('categoria')); status = texto(request.args.get('status'))
    termo = texto(request.args.get('q')); data_de = texto(request.args.get('data_de')); data_ate = texto(request.args.get('data_ate'))
    sql = '''
        SELECT o.id_ocorrencia, o.categoria, o.tipo_ocorrencia, o.descricao,
               o.arquivo, o.data_da_criacao, o.data_inicio_oc, o.data_fim_oc,
               o.hora_saida, o.quem_busca, o.motivo_rejeicao, o.resposta_gestao,
               o.registrado, o.saida_confirmada, o.data_saida_confirmada,
               a.matricula AS aluno_matricula, a.nome AS aluno_nome, a.turma AS aluno_turma,
               r.nome AS responsavel_nome, o.id_usuario_aprovador,
               ua.nivel_acesso AS aprovador_nivel, COALESCE(c.nome, ad.nome, ua.email) AS aprovador_nome
        FROM Ocorrencia o
        LEFT JOIN (SELECT id_ocorrencia, MIN(matricula) AS matricula FROM Ocorrencia_Aluno GROUP BY id_ocorrencia) oa ON oa.id_ocorrencia = o.id_ocorrencia
        LEFT JOIN Aluno a ON a.matricula = oa.matricula
        LEFT JOIN Responsavel r ON r.id_responsavel = o.id_responsavel
        LEFT JOIN Usuario ua ON ua.id_usuario = o.id_usuario_aprovador
        LEFT JOIN Coordenador c ON c.id_usuario = ua.id_usuario
        LEFT JOIN Administrador ad ON ad.id_usuario = ua.id_usuario
        WHERE 1=1
    '''
    params=[]
    if categoria in ('Atestado','Liberacao'): sql += ' AND o.categoria = %s'; params.append(categoria)
    if status == 'pendente': sql += ' AND o.registrado = FALSE AND o.motivo_rejeicao IS NULL'
    elif status == 'aprovado': sql += ' AND o.registrado = TRUE'
    elif status == 'rejeitado': sql += ' AND o.registrado = FALSE AND o.motivo_rejeicao IS NOT NULL'
    if termo:
        sql += ' AND (a.nome LIKE %s OR a.matricula LIKE %s OR a.turma LIKE %s)'; busca=f'%{termo}%'; params += [busca,busca,busca]
    if data_de: sql += ' AND DATE(o.data_da_criacao) >= %s'; params.append(data_de)
    if data_ate: sql += ' AND DATE(o.data_da_criacao) <= %s'; params.append(data_ate)
    sql += ' ORDER BY o.data_da_criacao DESC'
    registros=fetch_all(sql,tuple(params)); mapa={3:'Gestão',5:'Administrador',1:'Responsável',2:'Professor',4:'Porteiro'}
    for item in registros:
        item['aprovador_perfil']=mapa.get(item.get('aprovador_nivel'))
        item['status']='aprovado' if item.get('registrado') else ('rejeitado' if item.get('motivo_rejeicao') else 'pendente')
    return jsonify(registros)


@admin_bp.route('/turmas/<string:codigo>/grade', methods=['GET'])
@papel_obrigatorio('administrador')
def grade_turma(codigo):
    codigo = texto(codigo).upper()
    turma = fetch_one('SELECT codigo FROM Turma WHERE codigo = %s', (codigo,))
    if not turma:
        return jsonify({'erro': 'Turma não encontrada.'}), 404
    return jsonify(fetch_all('''
        SELECT h.id_horario, h.turma, h.dia_da_semana, h.hr_inicio, h.hr_final,
               h.data_inicio_vigencia, h.data_fim_vigencia,
               m.id_materia, m.nome AS materia,
               h.matricula_professor, p.nome AS professor
        FROM Horario h
        JOIN Materia m ON m.id_materia = h.id_materia
        LEFT JOIN Professor p ON p.matricula = h.matricula_professor
        WHERE h.turma = %s
        ORDER BY h.data_inicio_vigencia DESC,
          FIELD(h.dia_da_semana,'Segunda','Terca','Quarta','Quinta','Sexta','Sabado','Domingo'),
          h.hr_inicio
    ''', (codigo,)))


@admin_bp.route('/horarios/<int:id_horario>/professor', methods=['PUT'])
@papel_obrigatorio('administrador')
def vincular_professor_horario(id_horario):
    dados = request.get_json() or {}
    valor = dados.get('matricula_professor')

    if valor in (None, '', 'null'):
        matricula_professor = None
    else:
        try:
            matricula_professor = int(valor)
        except (TypeError, ValueError):
            return jsonify({'erro': 'Professor inválido.'}), 400

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute('''
            SELECT h.id_horario, h.id_materia, h.data_inicio_vigencia, h.data_fim_vigencia,
                   m.nome AS materia
            FROM Horario h
            JOIN Materia m ON m.id_materia = h.id_materia
            WHERE h.id_horario = %s
        ''', (id_horario,))
        horario = cur.fetchone()
        if not horario:
            return jsonify({'erro': 'Horário não encontrado.'}), 404

        # Grades cuja vigência já acabou são histórico e não devem ser reescritas.
        fim_vigencia = horario[3]
        if fim_vigencia and fim_vigencia < date.today():
            return jsonify({
                'erro': 'Esta aula pertence a uma grade histórica encerrada e não pode ter o professor alterado.'
            }), 409

        if matricula_professor is not None:
            cur.execute('''
                SELECT p.matricula
                FROM Professor p
                JOIN Usuario u ON u.id_usuario = p.id_usuario
                JOIN Professor_Materia pm ON pm.matricula_professor = p.matricula
                WHERE p.matricula = %s AND pm.id_materia = %s AND u.ativo = TRUE
            ''', (matricula_professor, horario[1]))
            if not cur.fetchone():
                return jsonify({
                    'erro': 'Esse professor não está ativo ou não está vinculado à matéria desta aula.'
                }), 409

        cur.execute(
            'UPDATE Horario SET matricula_professor = %s WHERE id_horario = %s',
            (matricula_professor, id_horario)
        )
        conn.commit()
        return jsonify({
            'mensagem': 'Professor da aula atualizado com sucesso.' if matricula_professor else 'Professor removido desta aula.'
        })
    except IntegrityError as exc:
        conn.rollback()
        return jsonify({'erro': erro_integridade(exc)}), 409
    finally:
        cur.close()
        conn.close()


@admin_bp.route('/turmas/<string:codigo>', methods=['PUT'])
@papel_obrigatorio('administrador')
def editar_turma(codigo):
    codigo = texto(codigo).upper()
    dados = request.get_json(silent=True) or {}
    novo_codigo = texto(dados.get('codigo')).upper()
    if not novo_codigo or len(novo_codigo) > 10:
        return jsonify({'erro': 'Informe um código de turma com até 10 caracteres.'}), 400
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute('SELECT codigo FROM Turma WHERE codigo=%s', (codigo,))
        if not cur.fetchone():
            return jsonify({'erro': 'Turma não encontrada.'}), 404
        if novo_codigo != codigo:
            cur.execute('SELECT codigo FROM Turma WHERE codigo=%s', (novo_codigo,))
            if cur.fetchone():
                return jsonify({'erro': 'Já existe uma turma com esse código.'}), 409
            # As FKs de Aluno e Horario usam ON UPDATE CASCADE, preservando todos os vínculos.
            cur.execute('UPDATE Turma SET codigo=%s WHERE codigo=%s', (novo_codigo, codigo))
        conn.commit()
        return jsonify({'mensagem': 'Turma atualizada com sucesso.', 'codigo': novo_codigo})
    except IntegrityError as exc:
        conn.rollback(); return jsonify({'erro': erro_integridade(exc)}), 409
    finally:
        cur.close(); conn.close()


@admin_bp.route('/turmas/<string:codigo>/grade', methods=['POST'])
@papel_obrigatorio('administrador')
def atualizar_grade_turma(codigo):
    codigo = texto(codigo).upper()
    arquivo = request.files.get('planilha')
    inicio_txt = texto(request.form.get('data_inicio_vigencia'))
    if not arquivo or not arquivo.filename or not arquivo.filename.lower().endswith('.xlsx'):
        return jsonify({'erro': 'Selecione uma planilha .xlsx válida.'}), 400
    try:
        inicio = datetime.strptime(inicio_txt, '%Y-%m-%d').date() if inicio_txt else date.today()
    except ValueError:
        return jsonify({'erro': 'Data de início da nova grade inválida.'}), 400
    if inicio.year != date.today().year:
        return jsonify({'erro': 'A nova grade deve iniciar no ano vigente.'}), 400
    try:
        aulas = extrair_grade_planilha(arquivo.stream)
    except ValueError as exc:
        return jsonify({'erro': str(exc)}), 400

    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute('SELECT codigo FROM Turma WHERE codigo=%s', (codigo,))
        if not cur.fetchone():
            return jsonify({'erro': 'Turma não encontrada.'}), 404

        # Uma nova vigência nunca apaga a grade histórica. Ela encerra a anterior
        # no dia imediatamente anterior e cria novos horários a partir da data informada.
        cur.execute('SELECT MAX(data_inicio_vigencia) FROM Horario WHERE turma=%s', (codigo,))
        ultima = cur.fetchone()[0]
        if ultima and inicio <= ultima:
            return jsonify({'erro': f'A nova grade deve começar depois da vigência mais recente ({ultima.strftime("%d/%m/%Y")}).'}), 409

        fim_anterior = inicio - timedelta(days=1)
        cur.execute('''
            UPDATE Horario SET data_fim_vigencia=%s
            WHERE turma=%s AND data_inicio_vigencia < %s
              AND (data_fim_vigencia IS NULL OR data_fim_vigencia >= %s)
        ''', (fim_anterior, codigo, inicio, inicio))

        fim_nova = date(inicio.year, 12, 31)
        materias = set(); sem_professor = 0
        for aula in aulas:
            id_materia = materia_id_por_nome(cur, aula['materia']); materias.add(id_materia)
            professor = professor_unico_da_materia(cur, id_materia)
            if professor is None: sem_professor += 1
            cur.execute('''
                INSERT INTO Horario
                (turma,id_materia,matricula_professor,dia_da_semana,hr_inicio,hr_final,
                 data_inicio_vigencia,data_fim_vigencia,id_usuario_cadastro)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ''', (codigo,id_materia,professor,aula['dia'],aula['inicio'],aula['fim'],inicio,fim_nova,g.usuario['id_usuario']))
        conn.commit()
        return jsonify({'mensagem':'Nova grade importada com sucesso. A grade anterior foi preservada no histórico.',
                        'aulas_importadas':len(aulas),'materias_identificadas':len(materias),'aulas_sem_professor':sem_professor}), 201
    except IntegrityError as exc:
        conn.rollback(); return jsonify({'erro': erro_integridade(exc)}), 409
    finally:
        cur.close(); conn.close()
